#!/usr/bin/env python3
"""Export every sheet of the master workbook to data/sheets_csv/ as a flat CSV.

The workbook is the source of truth; the CSVs are generated from it so that git
can diff them line by line and so that GitHub can render and anchor to rows.
Run this after any edit to the workbook, then commit both.

    python3 scripts/07_export_sheets.py
"""
import csv
import pathlib

import openpyxl

WORKBOOK = pathlib.Path("data/history_engine_master_2026-08.xlsx")
OUTDIR = pathlib.Path("data/sheets_csv")


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"workbook not found: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    OUTDIR.mkdir(parents=True, exist_ok=True)
    written = []
    for ws in wb.worksheets:
        path = OUTDIR / f"{ws.title}.csv"
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])
        written.append((ws.title, ws.max_row))
    for title, rows in written:
        print(f"{title:<22}{rows:>6} rows")
    print(f"\n{len(written)} sheets exported to {OUTDIR}/")


if __name__ == "__main__":
    main()
